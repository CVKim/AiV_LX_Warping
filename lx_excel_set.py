import os
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Border, Side, Alignment
from collections import defaultdict
from zipfile import BadZipFile
from PIL import Image
import tempfile

def load_excel(filepath, sheet_name='LX_4th_수평선_PoC'):
    """
    엑셀 파일을 로드하거나 새로 생성합니다.
    """
    if os.path.exists(filepath):
        try:
            wb = load_workbook(filepath)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"기존 엑셀 파일 '{filepath}'의 시트 '{sheet_name}'을 엽니다.")
            else:
                ws = wb.create_sheet(sheet_name)
                print(f"엑셀 파일 '{filepath}'에 시트 '{sheet_name}'을 새로 생성합니다.")
        except BadZipFile:
            print(f"선택한 파일 '{filepath}'은 유효한 Excel 파일이 아닙니다.")
            print("새 엑셀 파일을 생성합니다.")
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            headers = ["No.", "파일명", "영상 원본", "왜곡 보정 및 나사 측정 이미지"]
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        print(f"새 엑셀 파일 '{filepath}'과 시트 '{sheet_name}'을 생성합니다.")
        headers = ["No.", "파일명", "영상 원본", "왜곡 보정 및 나사 측정 이미지"]
        ws.append(headers)
    
    # 헤더가 이미 있는지 확인하고, 없으면 추가
    if ws.max_row == 1 and ws.max_column >= 4:
        existing_headers = [cell.value for cell in ws[1][:4]]
        expected_headers = ["No.", "파일명", "영상 원본", "왜곡 보정 및 나사 측정 이미지"]
        if existing_headers != expected_headers:
            ws.delete_rows(1)
            ws.append(expected_headers)
    else:
        headers = ["No.", "파일명", "영상 원본", "왜곡 보정 및 나사 측정 이미지"]
        ws.append(headers)
    
    return wb, ws

def insert_image_into_excel(ws, img_path, cell, resize=None, temp_files=None):
    """
    이미지를 엑셀 셀에 삽입합니다.
    - resize: (width, height) 튜플로 이미지 리사이즈. None이면 원본 사이즈 유지.
    """
    try:
        # 이미지 열기
        pil_image = Image.open(img_path)
        print(f"Processing image: {img_path}")
        print(f"Original size: {pil_image.size}")
        
        # 이미지 리사이즈
        if resize:
            try:
                # Pillow 10.0.0 이상
                pil_image = pil_image.resize(resize, Image.Resampling.LANCZOS)
            except AttributeError:
                # Pillow 9.x 이하
                pil_image = pil_image.resize(resize, Image.LANCZOS)
            print(f"Resized size: {pil_image.size}")
        
        # 이미지의 최소 크기 확인
        if pil_image.width < 1 or pil_image.height < 1:
            print(f"이미지 크기가 너무 작아 삽입할 수 없습니다: {img_path}")
            return
        
        # PNG 형식으로 저장하여 압축 방지 (원본 품질 유지)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
            pil_image.save(tmp.name, "PNG")
            tmp_path = tmp.name
        
        img = XLImage(tmp_path)
        img.anchor = cell
        ws.add_image(img)
        
        # 임시 파일 경로를 리스트에 추가하여 나중에 삭제
        if temp_files is not None:
            temp_files.append(tmp_path)
        
    except Exception as e:
        print(f"이미지 삽입 중 오류 발생 ({img_path}): {e}")

def apply_borders(ws, start_row, end_row, start_col, end_col):
    """
    지정된 범위의 셀에 테두리를 적용합니다.
    모든 셀에 얇은 테두리를 적용하고, 외곽에는 굵은 테두리를 추가합니다.
    """
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="thick", color="000000")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)
    
    # 모든 셀에 얇은 테두리와 가운데 정렬 적용
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 외곽 셀에 굵은 테두리 적용
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if cell.row == start_row:
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=thick,
                    bottom=cell.border.bottom
                )
            if cell.row == end_row:
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=thick
                )
            if cell.column == start_col:
                cell.border = Border(
                    left=thick,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
            if cell.column == end_col:
                cell.border = Border(
                    left=cell.border.left,
                    right=thick,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )

def main():
    # 1. 엑셀 파일 경로 지정 - 현재 디렉토리에 저장
    excel_filename = "LX_PoC.xlsx"
    excel_path = os.path.join(os.getcwd(), excel_filename)
    print(f"엑셀 파일 경로: {excel_path}")
    
    # 2. 엑셀 파일 로드 또는 생성
    wb, ws = load_excel(excel_path, sheet_name='LX_4th_수평선_PoC')
    
    # 3. 이미지 디렉토리 지정
    original_image_dir = r"\\192.168.10.74\01.Data\01.Image\LX\2_DATA_SET\241212\시공목_나사_측정 데이터"
    annotated_image_dir = r"\\192.168.10.74\01.Data\01.Image\LX\2_DATA_SET\241212\PoC_Data\수평선_검출 결과"
    
    # 4. 이미지 디렉토리 존재 여부 확인
    if not os.path.exists(original_image_dir):
        print(f"원본 이미지 디렉토리가 존재하지 않습니다: {original_image_dir}")
        return
    if not os.path.exists(annotated_image_dir):
        print(f"측정 이미지 디렉토리가 존재하지 않습니다: {annotated_image_dir}")
        return
    
    # 5. 이미지 파일 목록 가져오기
    # 원본 이미지 파일
    original_image_files = os.listdir(original_image_dir)
    original_image_files = [f for f in original_image_files if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
    
    # 측정 이미지 파일
    annotated_image_files = os.listdir(annotated_image_dir)
    annotated_image_files = [f for f in annotated_image_files if f.lower().endswith('_warped_expanded.jpg')]
    
    if not original_image_files:
        print(f"원본 이미지 디렉토리에 JPG/JPEG/PNG 이미지가 없습니다: {original_image_dir}")
    
    if not annotated_image_files:
        print(f"측정 이미지 디렉토리에 '_warped_expanded.jpg'가 포함된 이미지가 없습니다: {annotated_image_dir}")
    
    # 6. 파일명 기준으로 그룹화
    file_groups = defaultdict(dict)
    
    # 원본 이미지 그룹화
    for f in original_image_files:
        base_name, ext = os.path.splitext(f)
        file_groups[base_name]['original'] = os.path.join(original_image_dir, f)
    
    # 측정 이미지 그룹화
    for f in annotated_image_files:
        if f.lower().endswith('_warped_expanded.jpg'):
            base_name = f[:-len('_warped_expanded.jpg')]
            file_groups[base_name]['annotated'] = os.path.join(annotated_image_dir, f)
    
    # 7. 이미지 삽입
    temp_files = []  # 임시 파일 경로 저장 리스트
    current_row = ws.max_row  # 데이터 시작 행
    
    for base_name, paths in file_groups.items():
        # 'original'과 'annotated' 모두 존재하는 경우에만 삽입
        if 'original' in paths and 'annotated' in paths:
            # Verify that both image paths exist and are accessible
            original_img_path = paths['original']
            annotated_img_path = paths['annotated']
            
            if not os.path.exists(original_img_path):
                print(f"원본 이미지 파일이 존재하지 않습니다: {original_img_path}")
                continue
            if not os.path.exists(annotated_img_path):
                print(f"왜곡 보정 이미지 파일이 존재하지 않습니다: {annotated_img_path}")
                continue
            
            # Open images to check size
            try:
                pil_original = Image.open(original_img_path)
                pil_annotated = Image.open(annotated_img_path)
                print(f"Original Image Size: {pil_original.size}")
                print(f"Annotated Image Size: {pil_annotated.size}")
            except Exception as e:
                print(f"이미지 파일 열기 중 오류 발생 ({base_name}): {e}")
                continue
            
            # Only proceed if both images have valid sizes
            if pil_original.width < 1 or pil_original.height < 1:
                print(f"원본 이미지 크기가 너무 작아 삽입할 수 없습니다: {original_img_path}")
                continue
            if pil_annotated.width < 1 or pil_annotated.height < 1:
                print(f"왜곡 보정 이미지 크기가 너무 작아 삽입할 수 없습니다: {annotated_img_path}")
                continue
            
            # Increment row
            current_row += 1
            ws.cell(row=current_row, column=1, value=current_row - 1)  # No.
            ws.cell(row=current_row, column=2, value=base_name)  # 파일명
            
            # 영상 원본 삽입 (C열) - 150x150 픽셀로 리사이즈
            insert_image_into_excel(ws, original_img_path, f"C{current_row}", resize=(150, 150), temp_files=temp_files)
            
            # 왜곡 보정 및 나사 측정 이미지 삽입 (D열) - 너비 50으로 리사이즈, 높이는 비율에 맞게 조정
            try:
                pil_image = Image.open(annotated_img_path)
                if pil_image.width == 0:
                    print(f"이미지 너비가 0인 파일을 건너뜁니다: {annotated_img_path}")
                    continue
                width_percent = (50 / float(pil_image.size[0]))
                new_height = int((float(pil_image.size[1]) * float(width_percent)))
                if new_height < 1:
                    new_height = 1  # 최소 높이 설정
                # Ensure new_width is exactly 50
                new_width = 50
                insert_image_into_excel(ws, annotated_img_path, f"D{current_row}", resize=(new_width, new_height), temp_files=temp_files)
            except Exception as e:
                print(f"왜곡 보정 이미지 처리 중 오류 발생 ({annotated_img_path}): {e}")
                continue
        else:
            # 'annotated' 이미지가 없는 경우 건너뜀
            if 'annotated' not in paths:
                print(f"파일 '{base_name}'의 측정 이미지가 없습니다. 해당 파일을 엑셀에 기록하지 않습니다.")
            if 'original' not in paths:
                print(f"파일 '{base_name}'의 원본 이미지가 없습니다. 해당 파일을 엑셀에 기록하지 않습니다.")
    
    # 8. 엑셀 셀 크기 조정
    # 행 높이를 150으로 설정
    for row in range(2, current_row + 1):
        ws.row_dimensions[row].height = 150  # 포인트 단위
    
    # 열 너비를 설정
    ws.column_dimensions['A'].width = 10  # No. 열
    ws.column_dimensions['B'].width = 30  # 파일명 열
    ws.column_dimensions['C'].width = 30  # 영상 원본 열
    ws.column_dimensions['D'].width = 50   # 왜곡 보정 및 나사 측정 이미지 열
    
    # 열 너비 강제 설정 (ensure D column is at least 50)
    if ws.column_dimensions['D'].width < 50:
        ws.column_dimensions['D'].width = 50
    
    # 9. 테두리 적용 (A-D 열)
    apply_borders(ws, start_row=1, end_row=current_row, start_col=1, end_col=4)
    
    # 10. 엑셀 파일 저장
    try:
        wb.save(excel_path)
        print(f"엑셀 파일이 저장되었습니다: {excel_path}")
    except Exception as e:
        print(f"엑셀 파일 저장 중 오류 발생: {e}")
    
    # 11. 임시 파일 삭제
    for tmp_path in temp_files:
        try:
            os.remove(tmp_path)
        except Exception as e:
            print(f"임시 파일 삭제 중 오류 발생: {e}")

if __name__ == "__main__":
    main()
